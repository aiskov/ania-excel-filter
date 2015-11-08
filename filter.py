from openpyxl import load_workbook, Workbook
import utils
import csv
from settings import *

print('Source file opening')

# Open origin
wb_origin = load_workbook(filename = origin_filename)
ws_origin = wb_origin[origin_sheetname]

print('Source file opened')

# Check groups condition in case if exists
if not group_condition:
    selected_groups = None
else:
    selected_groups = []

    # Group rows 
    groups = dict()

    print('Starting grouping')
    for row in rows_range:
        group_value = ws_origin['%s%s' % (group_by, row)].value

    if group_value not in groups:
        groups[group_value] = []

    groups[group_value].append(row)

print('File grouped to %s by column %s.' % (len(groups), group_by))

# Check group condition
print('Starting group filtering')
for key, rows in groups.iteritems():
    if group_condition(ws_origin, rows):
        selected_groups.append(key)

print('%s groups found that match the conditions.' % len(selected_groups))

# Copy 
if target_filename.endswith('xlsx'):
    target_row = 1

    # Open target
    wb_target = Workbook()
    ws_target = wb_target.active
    ws_target.title = target_sheetname

    print('Start copying')
    for row in rows_range:
        check = not row_condition or row_condition(ws_origin, row)
        check = check and (not group_condition or ws_origin['%s%s' % (group_by, row)].value in selected_groups)

        if check:
            for column in columns:
                ws_target['%s%s' % (column, target_row)] = ws_origin['%s%s' % (column, row)].value

            target_row += 1

    # Save target file
    wb_target.save(filename = target_filename)
else:
    with open(target_filename, 'wb') as f:
        wb_target = csv.writer(f)

    print('Start copying')
    for row in rows_range:
        check = not row_condition or row_condition(ws_origin, row)
        check = check and (not group_condition or ws_origin['%s%s' % (group_by, row)].value in selected_groups)

        if check:
            row = []
            for column in columns:
                row.append(ws_origin['%s%s' % (column, row)].value)

        wb_target.writerow(row)


# Print report and wait input
raw_input("Press Enter to continue...")
